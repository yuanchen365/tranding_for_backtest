from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter


def export_excel(
    output_path: Path,
    strategies: List[str],
    trade_log: pd.DataFrame,
    monthly_tables: Dict[str, pd.DataFrame],
    yearly_tables: Dict[str, pd.DataFrame],
    plot_paths: Dict[str, Dict[str, Path]],
    equity_path: Optional[Path],
    performance_df: pd.DataFrame,
    n: int,
    l: int,
    embed_images: bool,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for strategy_label in strategies:
            df_strategy = trade_log[trade_log["Strategy"] == strategy_label].copy()
            sheet_name = strategy_label.replace(" ", "_")[:31] or "Strategy"
            if df_strategy.empty:
                df_strategy = pd.DataFrame(columns=trade_log.columns)
            df_strategy.to_excel(writer, index=False, sheet_name=sheet_name)

            ws = writer.sheets[sheet_name]
            try:
                ws.freeze_panes = "A2"
            except Exception:
                pass

            month_tbl = monthly_tables.get(strategy_label, pd.DataFrame(columns=["Year", "Month", "勝率百分比", "出場筆數", "年月"]))
            year_tbl = yearly_tables.get(strategy_label, pd.DataFrame(columns=["Year", "年淨利"]))

            start_row = len(df_strategy) + 2
            info_left = pd.DataFrame({f"{sheet_name} 月勝率（%）": [f"N={n}, L={l}"]})
            info_left.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row, startcol=0)
            month_tbl.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row + 1, startcol=0)

            year_start_col = max(3, month_tbl.shape[1]) + 2
            info_right = pd.DataFrame({f"{sheet_name} 年淨利": [f"N={n}, L={l}"]})
            info_right.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row, startcol=year_start_col)
            year_tbl.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row + 1, startcol=year_start_col)

            if embed_images:
                start_offset = start_row + max(len(month_tbl) + 3, len(year_tbl) + 3) + 2
                imgs = plot_paths.get(strategy_label, {})
                monthly_img = imgs.get("monthly")
                if monthly_img and monthly_img.exists():
                    ws.add_image(XLImage(str(monthly_img)), f"A{start_offset}")
                yearly_img = imgs.get("yearly")
                if yearly_img and yearly_img.exists():
                    col_letter = get_column_letter(year_start_col + 1)
                    ws.add_image(XLImage(str(yearly_img)), f"{col_letter}{start_offset}")
            else:
                ws.cell(row=start_row, column=1, value="（已停用圖片嵌入）")

        # SUMMARY sheet
        perf_df = performance_df.copy()
        if not perf_df.empty:
            perf_df["N"] = n
            perf_df["L"] = l
        perf_df.to_excel(writer, sheet_name="SUMMARY")
        summary_ws = writer.sheets["SUMMARY"]

        if embed_images and equity_path and equity_path.exists():
            anchor_row = summary_ws.max_row + 2
            summary_ws.add_image(XLImage(str(equity_path)), f"A{anchor_row}")
        elif not embed_images:
            summary_ws.cell(row=1, column=perf_df.shape[1] + 2, value="（已停用圖片嵌入）")
